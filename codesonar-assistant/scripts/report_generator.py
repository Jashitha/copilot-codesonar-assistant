"""
report_generator.py

Builds the two-sheet CodeSonar tracker report (Summary + Details) using openpyxl.

Public API:
    generate_summary_sheet(workbook, dataframe)
    generate_details_sheet(workbook, dataframe)
    save_tracker_report(df, output_file)
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=12)
CENTER_ALIGN = Alignment(horizontal="center")

# Exact column order required for the Details sheet.
DETAILS_COLUMNS = [
    "score",
    "id",
    "class",
    "significance",
    "file",
    "line number",
    "procedure",
    "priority",
    "state",
    "finding",
    "owner",
    "reviewer",
    "url",
]

TOP_N = 5


def _blank_owner_mask(series: pd.Series) -> pd.Series:
    """True for rows whose owner is blank/NaN/'unassigned'/'none'."""
    normalized = series.astype(str).str.strip().str.lower()
    return normalized.isin(["", "nan", "unassigned", "none"])


def _metric_rows(df: pd.DataFrame) -> list[tuple[str, int]]:
    """Overall Metrics: Total Issues, Pending, Done, HB_PRIO_1, HB_PRIO_2, Owners, Reviewers."""
    status_col = "Status" if "Status" in df.columns else "state"
    owner_col = "Owner" if "Owner" in df.columns else "owner"
    reviewer_col = "Reviewer" if "Reviewer" in df.columns else "reviewer"

    status = (
        df[status_col].astype(str).str.lower()
        if status_col in df.columns
        else pd.Series(dtype=str)
    )
    priority = df["priority"] if "priority" in df.columns else pd.Series(dtype=str)

    total = len(df)
    pending = int((status == "pending").sum())
    done = int((status == "done").sum())
    hb1 = int((priority == "HB_PRIO_1").sum())
    hb2 = int((priority == "HB_PRIO_2").sum())

    if owner_col in df.columns:
        owners = df.loc[~_blank_owner_mask(df[owner_col]), owner_col].astype(str).str.strip()
        owner_count = owners.nunique()
    else:
        owner_count = 0

    if reviewer_col in df.columns:
        reviewers = df.loc[~_blank_owner_mask(df[reviewer_col]), reviewer_col].astype(str).str.strip()
        reviewer_count = reviewers.nunique()
    else:
        reviewer_count = 0

    return [
        ("Total Issues", total),
        ("Pending", pending),
        ("Done", done),
        ("HB_PRIO_1", hb1),
        ("HB_PRIO_2", hb2),
        ("Owners", owner_count),
        ("Reviewers", reviewer_count),
    ]


def _top_files(df: pd.DataFrame, limit: int = TOP_N) -> list[tuple[str, int]]:
    """Top N files with the highest issue count, sorted descending."""
    if "file" not in df.columns:
        return []
    counts = (
        df.groupby("file").size().sort_values(ascending=False).head(limit)
    )
    return list(counts.items())


def _class_distribution(df: pd.DataFrame) -> list[tuple[str, int]]:
    """Full issue-class distribution, sorted descending."""
    if "class" not in df.columns:
        return []
    counts = df.groupby("class").size().sort_values(ascending=False)
    return list(counts.items())


def _workload(df: pd.DataFrame, primary_col: str, fallback_col: str) -> list[tuple[str, int]]:
    """Issue count per assignee for the given column, sorted descending.

    Blank/NaN/'unassigned'/'none' values collapse into a single 'Unassigned' row so the
    split across real owners/reviewers is easy to read.
    """
    col = primary_col if primary_col in df.columns else fallback_col
    if col not in df.columns:
        return []

    values = df[col].astype(str).str.strip()
    values = values.mask(_blank_owner_mask(values), "Unassigned")
    counts = values.value_counts()
    return list(counts.items())


def _write_section(
    ws: Worksheet,
    start_row: int,
    title: str,
    headers: list[str],
    rows: list[tuple],
) -> int:
    """Write a titled, formatted table starting at start_row. Returns the next free row."""
    row = start_row

    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    row += 1

    header_row = row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1

    for record in rows:
        for col_idx, value in enumerate(record, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx > 1:
                cell.alignment = CENTER_ALIGN
        row += 1

    # Blank spacer rows between sections.
    return row + 2


def _autosize_columns(ws: Worksheet, padding: int = 2) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))

    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width + padding


def generate_summary_sheet(workbook: Workbook, dataframe: pd.DataFrame) -> Worksheet:
    """Build the 'Summary' worksheet: Overall Metrics, Top Files, Top Issue Classes,
    and the Complete Issue Class Distribution."""
    if "Summary" in workbook.sheetnames:
        del workbook["Summary"]
    ws = workbook.create_sheet("Summary")

    class_dist = _class_distribution(dataframe)

    row = 1
    row = _write_section(ws, row, "Overall Metrics", ["Metric", "Value"], _metric_rows(dataframe))
    row = _write_section(ws, row, "Top Files", ["Top Files", "Issues"], _top_files(dataframe))
    row = _write_section(
        ws, row, "Top Issue Classes", ["Top Issue Class", "Issues"], class_dist[:TOP_N]
    )
    row = _write_section(
        ws, row, "Owner Workload", ["Owner", "Issues"], _workload(dataframe, "Owner", "owner")
    )
    row = _write_section(
        ws, row, "Reviewer Workload", ["Reviewer", "Issues"], _workload(dataframe, "Reviewer", "reviewer")
    )
    row = _write_section(
        ws, row, "Complete Issue Class Distribution", ["Class", "Issues"], class_dist
    )

    _autosize_columns(ws)
    ws.freeze_panes = "A2"
    return ws


def generate_details_sheet(workbook: Workbook, dataframe: pd.DataFrame) -> Worksheet:
    """Build the 'Details' worksheet with every issue row in the required column order."""
    if "Details" in workbook.sheetnames:
        del workbook["Details"]
    ws = workbook.create_sheet("Details")

    details_df = dataframe.rename(
        columns={"Status": "state", "Owner": "owner", "Reviewer": "reviewer"}
    )
    ordered = [c for c in DETAILS_COLUMNS if c in details_df.columns]
    details_df = details_df[ordered]

    for col_idx, header in enumerate(details_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for row_idx, record in enumerate(details_df.itertuples(index=False, name=None), start=2):
        for col_idx, value in enumerate(record, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    _autosize_columns(ws)
    ws.freeze_panes = "A2"
    return ws


def save_tracker_report(df: pd.DataFrame, output_file: str | Path) -> Path:
    """Build the Summary and Details sheets and save them as a single workbook."""
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    generate_summary_sheet(workbook, df)
    generate_details_sheet(workbook, df)

    workbook.save(output_path)
    return output_path
